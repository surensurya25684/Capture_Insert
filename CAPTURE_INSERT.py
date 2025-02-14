import streamlit as st
import pdfplumber
import re
import io
from openpyxl import Workbook

st.set_page_config(page_title="AGM Results Extractor", layout="wide")
st.title("AGM Results Extractor and Formatter")
st.markdown("""
This app extracts proposal‐ and director election–related details from an AGM PDF document.
It creates an Excel file with two sheets:
• Proposal Sheet (for proposals)  
• Non-Proposal Sheet (for director election data)

Note: This is a sample implementation. Proposals and non-proposals in your document can vary.
You may need to adjust the regex patterns and parsing logic to match your document’s format.
""")

uploaded_file = st.file_uploader("Upload your AGM PDF Document", type=["pdf"])

def extract_pdf_text(pdf_file):
    text = ""
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    return text

def parse_agm_proposals(text):
    """
    Parses proposals from the section starting at a designated header.
    This implementation assumes the proposals are found under a section like:
    "Item 5.07. Submission of Matters to a Vote of Security Holders."
    
    Adjust the header string and regex patterns as needed.
    """
    proposals = []
    try:
        # Change this header if your document uses a different marker
        block = text.split("Item 5.07. Submission of Matters to a Vote of Security Holders.")[1]
    except IndexError:
        st.error("Could not find the proposals section in the document.")
        return proposals

    # Regex to capture proposals.
    # This pattern assumes a format such as "1. Proposal 1 – <Title>" followed by details.
    pattern = re.compile(
        r"(\d+)\.\s*Proposal\s*(\d+)\s*–\s*(.*?)\n(.*?)(?=\n\d+\.\s*Proposal|\nNo other matters)",
        re.DOTALL
    )
    matches = pattern.findall(block)
    for match in matches:
        proposal_index, proposal_number, title, details = match
        proposal_title = title.strip()
        # For Proposal 1 (director election), preserve the entire block.
        if proposal_number == "1":
            proposals.append({
                "proposal_number": proposal_number,
                "type": "director_election",
                "director_block": details.strip(),
                "proposal_title": proposal_title
            })
        else:
            # For proposals 2,3,4: get the first non-empty line as description.
            lines = [ln for ln in details.strip().splitlines() if ln.strip() != ""]
            if not lines:
                continue
            description = lines[0].strip()
            # Initialize vote fields
            vote_for = ""
            vote_against = ""
            vote_abstain = ""
            vote_broker = ""
            # For Proposal 4, extract separate vote values (One Year, Two Years, Three Years)
            vote_one_year = ""
            vote_two_years = ""
            vote_three_years = ""
            header_line = ""
            value_line = ""
            # Look for a line with vote headers – adjust the logic if your document varies
            for i, line in enumerate(lines[1:], start=1):
                if ("For" in line and "Against" in line) or ("One Year" in line):
                    header_line = line.strip()
                    if i+1 < len(lines):
                        value_line = lines[i+1].strip()
                    break
            if header_line and "One Year" in header_line:
                # Proposal 4 with separate vote fields
                values = value_line.split()
                if len(values) >= 5:
                    vote_one_year = values[0]
                    vote_two_years = values[1]
                    vote_three_years = values[2]
                    vote_abstain = values[3]
                    vote_broker = values[4]
                else:
                    vote_one_year = vote_two_years = vote_three_years = vote_abstain = vote_broker = ""
                vote_for = ""  # Not used for Proposal 4
            elif header_line:
                # For other proposals: expect order: For, Against, Abstain, Broker Non-Votes.
                values = value_line.split()
                vote_for = values[0] if len(values) > 0 else ""
                vote_against = values[1] if len(values) > 1 else ""
                vote_abstain = values[2] if len(values) > 2 else ""
                vote_broker = values[3] if len(values) > 3 else ""
            # Determine resolution outcome if possible
            try:
                if vote_against:
                    if int(vote_for.replace(',', '')) > int(vote_against.replace(',', '')):
                        outcome = f"Approved ({vote_for} > {vote_against})"
                    else:
                        outcome = f"Not Approved ({vote_for} > {vote_against})"
                else:
                    outcome = f"Approved ({vote_for} > )" if vote_for else ""
            except:
                outcome = ""
            proposal_dict = {
                "proposal_number": proposal_number,
                "type": "proposal",
                "proposal_title": proposal_title,
                "description": description,
                "Proposal Proxy Year": "2024",  # Adjust as needed
                "Resolution Outcome": outcome,
                "Vote Results - For": vote_for,
                "Vote Results - Against": vote_against,
                "Vote Results - Abstained": vote_abstain,
                "Vote Results - Withheld": "",  # Not provided
                "Vote Results - Broker Non-Votes": vote_broker,
                "Mgmt Proposal Category": "",
                "Proposal Vote Results Total": ""
            }
            # If Proposal 4 format is detected, add the separate vote fields.
            if header_line and "One Year" in header_line:
                proposal_dict["Vote Results - One Year"] = vote_one_year
                proposal_dict["Vote Results - Two Years"] = vote_two_years
                proposal_dict["Vote Results - Three Years"] = vote_three_years
            proposals.append(proposal_dict)
    return proposals

def parse_director_elections(director_block):
    """
    Parses the director election table from Proposal 1’s block.
    Expected table header: "Nominee For Against Abstain Broker Non-Votes"
    Adjust the splitting logic if your table uses a different format.
    """
    directors = []
    lines = director_block.splitlines()
    start = False
    for line in lines:
        if "Nominee" in line and "For" in line and "Against" in line:
            start = True
            continue
        if start:
            if line.strip() == "" or line.strip().startswith("Proposal"):
                break
            parts = re.split(r'\s{2,}', line.strip())
            if len(parts) >= 4:
                name = parts[0]
                vote_for = parts[1]
                vote_against = parts[2]
                vote_abstain = parts[3]
                vote_broker = parts[4] if len(parts) > 4 else ""
                directors.append({
                    "Director Election Year": "2024",
                    "Individual": name,
                    "Director Votes For": vote_for,
                    "Director Votes Against": vote_against,
                    "Director Votes Abstained": vote_abstain,
                    "Director Votes Withheld": "",
                    "Director Votes Broker-Non-Votes": vote_broker
                })
    return directors

def separate_proposals(all_proposals):
    proposals_list = []
    director_entries = []
    for item in all_proposals:
        if item["type"] == "director_election":
            director_entries.append(item)
        else:
            proposals_list.append(item)
    return proposals_list, director_entries

def create_excel(proposals, directors):
    wb = Workbook()
    # Proposal Sheet
    ws1 = wb.active
    ws1.title = "Proposal"
    row_idx = 1
    for prop in proposals:
        ws1.cell(row=row_idx, column=1, value="Proposal Proxy Year:")
        ws1.cell(row=row_idx, column=2, value=prop["Proposal Proxy Year"])
        row_idx += 1

        ws1.cell(row=row_idx, column=1, value="Resolution Outcome:")
        ws1.cell(row=row_idx, column=2, value=prop["Resolution Outcome"])
        row_idx += 1

        ws1.cell(row=row_idx, column=1, value="Proposal Text:")
        ws1.cell(row=row_idx, column=2, value=f'"{prop["proposal_title"]}: {prop["description"]}"')
        row_idx += 1

        ws1.cell(row=row_idx, column=1, value="Mgmt Proposal Category:")
        ws1.cell(row=row_idx, column=2, value=prop["Mgmt Proposal Category"])
        row_idx += 1

        # Check if separate vote fields exist (e.g., Proposal 4 format)
        if "Vote Results - One Year" in prop:
            ws1.cell(row=row_idx, column=1, value="Vote Results - One Year:")
            ws1.cell(row=row_idx, column=2, value=prop["Vote Results - One Year"])
            row_idx += 1

            ws1.cell(row=row_idx, column=1, value="Vote Results - Two Years:")
            ws1.cell(row=row_idx, column=2, value=prop["Vote Results - Two Years"])
            row_idx += 1

            ws1.cell(row=row_idx, column=1, value="Vote Results - Three Years:")
            ws1.cell(row=row_idx, column=2, value=prop["Vote Results - Three Years"])
            row_idx += 1
        else:
            ws1.cell(row=row_idx, column=1, value="Vote Results - For:")
            ws1.cell(row=row_idx, column=2, value=prop["Vote Results - For"])
            row_idx += 1

            ws1.cell(row=row_idx, column=1, value="Vote Results - Against:")
            ws1.cell(row=row_idx, column=2, value=prop["Vote Results - Against"])
            row_idx += 1

        ws1.cell(row=row_idx, column=1, value="Vote Results - Abstained:")
        ws1.cell(row=row_idx, column=2, value=prop["Vote Results - Abstained"])
        row_idx += 1

        ws1.cell(row=row_idx, column=1, value="Vote Results - Withheld:")
        ws1.cell(row=row_idx, column=2, value=prop["Vote Results - Withheld"])
        row_idx += 1

        ws1.cell(row=row_idx, column=1, value="Vote Results - Broker Non-Votes:")
        ws1.cell(row=row_idx, column=2, value=prop["Vote Results - Broker Non-Votes"])
        row_idx += 1

        ws1.cell(row=row_idx, column=1, value="Proposal Vote Results Total:")
        ws1.cell(row=row_idx, column=2, value=prop["Proposal Vote Results Total"])
        row_idx += 1

        # Separator row
        ws1.cell(row=row_idx, column=1, value="--------------------------------------------------")
        row_idx += 1

    # Non-Proposal Sheet (Director Elections)
    ws2 = wb.create_sheet(title="Non-Proposal")
    row_idx = 1
    for director in directors:
        ws2.cell(row=row_idx, column=1, value="Director Election Year:")
        ws2.cell(row=row_idx, column=2, value=director["Director Election Year"])
        row_idx += 1

        ws2.cell(row=row_idx, column=1, value="Individual:")
        ws2.cell(row=row_idx, column=2, value=director["Individual"])
        row_idx += 1

        ws2.cell(row=row_idx, column=1, value="Director Votes For:")
        ws2.cell(row=row_idx, column=2, value=director["Director Votes For"])
        row_idx += 1

        ws2.cell(row=row_idx, column=1, value="Director Votes Against:")
        ws2.cell(row=row_idx, column=2, value=director["Director Votes Against"])
        row_idx += 1

        ws2.cell(row=row_idx, column=1, value="Director Votes Abstained:")
        ws2.cell(row=row_idx, column=2, value=director["Director Votes Abstained"])
        row_idx += 1

        ws2.cell(row=row_idx, column=1, value="Director Votes Withheld:")
        ws2.cell(row=row_idx, column=2, value=director["Director Votes Withheld"])
        row_idx += 1

        ws2.cell(row=row_idx, column=1, value="Director Votes Broker-Non-Votes:")
        ws2.cell(row=row_idx, column=2, value=director["Director Votes Broker-Non-Votes"])
        row_idx += 1

        # Separator row
        ws2.cell(row=row_idx, column=1, value="--------------------------------------------------")
        row_idx += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

if uploaded_file is not None:
    pdf_text = extract_pdf_text(uploaded_file)
    all_proposals = parse_agm_proposals(pdf_text)
    proposals_list, director_entries = separate_proposals(all_proposals)
    
    # Parse director elections from Proposal 1 block(s)
    directors = []
    for entry in director_entries:
        directors.extend(parse_director_elections(entry.get("director_block", "")))
    
    st.write("### Extracted Proposal Data (Proposals 2, 3, 4)")
    if proposals_list:
        for idx, prop in enumerate(proposals_list, start=1):
            st.write(f"**Proposal {prop['proposal_number']}:** {prop['proposal_title']}")
            st.write(f"Description: {prop.get('description', '')}")
            if "Vote Results - One Year" in prop:
                st.write(f"Votes One Year: {prop['Vote Results - One Year']}, Two Years: {prop['Vote Results - Two Years']}, Three Years: {prop['Vote Results - Three Years']}")
            else:
                st.write(f"Votes For: {prop['Vote Results - For']}, Votes Against: {prop['Vote Results - Against']}")
            st.write(f"Abstained: {prop['Vote Results - Abstained']}, Broker Non-Votes: {prop['Vote Results - Broker Non-Votes']}")
            st.write(f"Outcome: {prop['Resolution Outcome']}")
            st.write("---")
    else:
        st.warning("No proposals (other than director election) found.")

    st.write("### Extracted Director Election Data (Proposal 1)")
    if directors:
        for idx, dir_data in enumerate(directors, start=1):
            st.write(f"**Director {idx}:** {dir_data['Individual']} - Votes For: {dir_data['Director Votes For']}, Against: {dir_data['Director Votes Against']}, Abstained: {dir_data['Director Votes Abstained']}, Broker Non-Votes: {dir_data['Director Votes Broker-Non-Votes']}")
            st.write("---")
    else:
        st.warning("No director election data found.")

    excel_file = create_excel(proposals_list, directors)
    st.download_button(
        label="Download Excel File",
        data=excel_file,
        file_name="AGM_Results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

